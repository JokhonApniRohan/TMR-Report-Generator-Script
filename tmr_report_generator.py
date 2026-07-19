"""
=============================================================================
TMR (Trade Marketing Representative) Report Generator
upay Sales Operations — with Windows GUI
=============================================================================

HOW TO RUN:
    Double-click this .py file, OR run: python tmr_report_generator.py

A graphical window will open. You will:
  1. Select one or more TMR Daily Activity Excel files (bulk select allowed)
  2. Select one or more TMR Datewise Summary Excel files (bulk select allowed)
  3. Select the DH Wise Target Excel file (one file, defines visit targets by DH)
  4. Choose where to save the output report
  5. Click "Generate Report"

=============================================================================
FILE DESCRIPTIONS:
  - TMR Daily Activity : Each row = one agent visit by a TMR. Used for
                         precise attendance evaluation (timing + duration).
  - TMR Datewise Summary: Each row = one TMR's daily summary (visit count,
                          working hours). Used for KPI calculations.
  - DH Wise Target File : Defines daily visit targets per Distribution House.
                          Columns: DH Code, Distributor House Name,
                                   Market type, Daily visit target
  - Region Wise Salary  : Defines monthly total salary per Region.
                          Columns: Region, Fixed Salary, Variable Salary,
                                   Total Salary, Salary Month Day Count,
                                   Daily Fixed Salary, Daily Variable Salary
  - Leave Status File   : Approved TMR leaves (optional).
                          Columns: SL, Wallet Number, TMR Name, DH Code,
                                   Leave Type, Leave Day, Leave Start Day,
                                   Leave End Day, Approver, Remarks
                          Friday leaves are automatically cancelled.
=============================================================================
ATTENDANCE RULES:
  A TMR is marked PRESENT (P) on a day if:
    • They completed ≥ 75% of their daily visit target
    • Only visits between 10:00 AM – 7:00 PM are counted
    • Only visits with duration between 3 min and 10 min are counted
  Otherwise marked ABSENT (A).
  Friday = weekly holiday (blank cell).
=============================================================================
"""

import os
import sys
import warnings
import threading
import traceback
from datetime import time
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── tkinter (standard library, ships with Python on Windows) ──
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from decimal import Decimal, ROUND_HALF_UP

warnings.filterwarnings("ignore")


# =============================================================================
# SECTION 1: CONSTANTS & CONFIGURATION
# =============================================================================

# Attendance validity thresholds
VISIT_MIN_SECONDS    = 3 * 60     # visits shorter than 3 min are invalid
VISIT_MAX_SECONDS    = 10 * 60    # visits longer than 10 min are invalid
BUSINESS_START       = time(10, 0, 0)   # field work starts at 10:00 AM
BUSINESS_END         = time(19, 0, 0)   # field work ends at  07:00 PM
MIN_VISIT_ACHIEVE    = 0.75        # 75% of daily target must be met for P

# Market hour KPI target (same for everyone)
MARKET_HOUR_TARGET_H = 8           # 8 hours per working day (KPI rule)

# Salary KPI weights (must sum to 1.0)
KPI_WEIGHT_STRIKE    = 0.35   # Strike Rate Achievement%
KPI_WEIGHT_FREQUENCY = 0.40   # Frequency Wise (Weekly Coverage) Achievement%
KPI_WEIGHT_MKT_HOUR  = 0.25   # Market Hour Achievement%

# Fixed charges applied to every TMR
INTERNET_CHARGE      = 300    # BDT per month (flat)
CASHOUT_CHARGE_PCT   = 0.01   # 1% of Total Salary

# upay brand palette (used in the Excel output)
C_BLUE   = "0054A5"
C_YELLOW = "FFD504"
C_WHITE  = "FFFFFF"
C_LIGHT  = "EEF3FB"   # alternating row tint
C_GREEN  = "00B050"   # Present cell
C_RED    = "C00000"   # Absent cell
C_GRAY   = "D6D6D6"   # Friday / no-data cell


# =============================================================================
# SECTION 2: UTILITY FUNCTIONS
# =============================================================================

def hms_to_seconds(hms_str):
    """
    Convert 'H:M:S' or 'H:M:S.ffffff' strings (from the Excel files) to
    an integer number of seconds. Returns 0 for any unreadable value.

    Examples:
        "8:4:42"          → 29082
        "0:13:40"         → 820
        "0:5:5.123456"    → 305
    """
    if pd.isna(hms_str):
        return 0
    s = str(hms_str).strip().split(".")[0]   # drop microseconds
    parts = s.split(":")
    if len(parts) == 3:
        try:
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
            return h * 3600 + m * 60 + sec
        except ValueError:
            return 0
    return 0


def seconds_to_hms(total_seconds):
    """
    Convert integer seconds back to a human-readable 'H:MM:SS' string.
    Used when writing Average Work Time into the report.
    """
    if not total_seconds or total_seconds <= 0:
        return "0:00:00"
    total_seconds = int(total_seconds)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h}:{m:02d}:{s:02d}"


def is_friday(ts):
    """
    Returns True if the given Timestamp/date is a Friday.
    In Bangladesh, Friday is the weekly public holiday.
    weekday() → 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
    """
    return pd.Timestamp(ts).weekday() == 4


def count_working_days(start, end):
    """
    Count non-Friday days between start and end dates (inclusive).
    These are the days TMRs are expected to work.
    """
    days = pd.date_range(start=start, end=end, freq="D")
    return sum(1 for d in days if d.weekday() != 4)


def count_fridays(start, end):
    """Count Fridays (weekly holidays) between start and end (inclusive)."""
    days = pd.date_range(start=start, end=end, freq="D")
    return sum(1 for d in days if d.weekday() == 4)


def normalize_wallet(wallet_str):
    """
    Strip leading zeros and whitespace from wallet numbers so that
    '01763489307' and '1763489307' both match in lookups.
    """
    return str(wallet_str).strip().lstrip("0")


def format_wallet_for_display(wallet_str):
    """
    Restore the leading zero for display in the output Excel
    (Bangladesh mobile numbers are 11 digits starting with 0).
    '1763489307' → '01763489307'
    """
    w = normalize_wallet(wallet_str)
    return ("0" + w) if len(w) == 10 else w


# =============================================================================
# SECTION 3: FILE LOADERS
# =============================================================================

def load_activity_files(file_paths, log):
    """
    Load and concatenate all TMR Daily Activity Excel files.

    Each file has one row per agent visit. Key columns used:
      - tmr_wallet     : identifies the TMR
      - checkin_time   : when visit started (used for business-hours filter)
      - checkout_time  : when visit ended
      - visit_time_h_m_s : visit duration as H:M:S string
      - created_at     : date of the visit record
      - dh_code        : Distribution House code (for target lookup)

    Returns a combined DataFrame with all visits across all uploaded files.
    """
    frames = []
    for fp in file_paths:
        log(f"  Loading activity: {os.path.basename(fp)}")
        df = pd.read_excel(fp, dtype={"tmr_wallet": str, "agent_wallet": str, "dh_code": str})
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    activity = pd.concat(frames, ignore_index=True)

    # Parse all datetime columns; coerce bad values to NaT
    activity["checkin_time"]  = pd.to_datetime(activity["checkin_time"],  errors="coerce")
    activity["checkout_time"] = pd.to_datetime(activity["checkout_time"], errors="coerce")
    activity["created_at"]    = pd.to_datetime(activity["created_at"],    errors="coerce")

    # Pre-compute visit duration in seconds from the H:M:S string column
    activity["visit_seconds"] = activity["visit_time_h_m_s"].apply(hms_to_seconds)

    # Extract a plain date from the timestamp (e.g. 2026-06-01 00:00:00)
    activity["date"] = activity["created_at"].dt.normalize()

    # Normalize wallet for consistent lookups
    activity["tmr_wallet"] = activity["tmr_wallet"].apply(normalize_wallet)

    log(f"  → {len(activity):,} visit rows | {activity['date'].nunique()} unique dates")
    return activity


def load_summary_files(file_paths, log):
    """
    Load and concatenate all TMR Datewise Summary Excel files.

    Each file has one row per TMR per day. Key columns used:
      - tmr_wallet         : identifies the TMR
      - date_              : the date of the summary row
      - agent_visit_count  : how many agents the TMR visited that day
      - working_time_h_m_s : total market hours that day
      - dh_code            : Distribution House code

    Returns a combined DataFrame with all daily summaries.
    """
    frames = []
    for fp in file_paths:
        log(f"  Loading summary:  {os.path.basename(fp)}")
        df = pd.read_excel(fp, dtype={"tmr_wallet": str, "dh_code": str})
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    summary = pd.concat(frames, ignore_index=True)

    # Normalize date column
    summary["date_"] = pd.to_datetime(summary["date_"], errors="coerce").dt.normalize()

    # Pre-compute working time in seconds
    summary["working_seconds"] = summary["working_time_h_m_s"].apply(hms_to_seconds)

    # Normalize wallet
    summary["tmr_wallet"] = summary["tmr_wallet"].apply(normalize_wallet)

    log(f"  → {len(summary):,} summary rows loaded")
    return summary


def load_dh_targets(file_path, log):
    """
    Load the DH Wise Target file.

    Expected columns (case-insensitive, flexible):
      DH Code               → unique code of the Distribution House
      Distributor House Name→ human-readable DH name
      Market type           → 'Metro' or 'Non metro'
      Daily visit target    → integer, e.g. 30, 35, or 40

    Returns a dict:  { normalized_dh_code → daily_visit_target_int }
    And a second dict: { normalized_dh_code → market_type_str }
    """
    log(f"  Loading DH targets: {os.path.basename(file_path)}")

    df = pd.read_excel(file_path, dtype={"DH Code": str})

    # ── Flexible column name matching (case-insensitive, strip spaces) ──
    df.columns = df.columns.str.strip()

    # Build a lowercase→actual map
    col_map = {c.lower().replace(" ", "_"): c for c in df.columns}

    # Find the columns we need
    def find_col(*candidates):
        for c in candidates:
            if c in col_map:
                return col_map[c]
        return None

    dh_code_col   = find_col("dh_code", "dhcode")
    target_col    = find_col("daily_visit_target", "daily_target", "visit_target")
    mtype_col     = find_col("market_type", "markettype", "type")

    if not dh_code_col or not target_col:
        raise ValueError(
            "DH Target file must have 'DH Code' and 'Daily visit target' columns. "
            f"Found columns: {list(df.columns)}"
        )

    dh_target_map = {}
    dh_mtype_map  = {}

    for _, row in df.iterrows():
        dh = str(row[dh_code_col]).strip()
        if not dh or dh.lower() == "nan":
            continue
        try:
            target = int(row[target_col])
        except (ValueError, TypeError):
            continue
        dh_target_map[dh] = target
        if mtype_col:
            dh_mtype_map[dh] = str(row[mtype_col]).strip()

    log(f"  → {len(dh_target_map)} DH target entries loaded")
    return dh_target_map, dh_mtype_map


def load_salary_file(file_path, report_end_date, log):
    """
    Load the Region Wise Salary file.

    Expected columns (row 1 = headers, no blank header row):
      Region               → region name (matches TMR region field)
      Total Salary         → monthly gross salary for the region
      Fixed Salary         → 70% of Total Salary  (or as defined)
      Variable Salary      → 30% of Total Salary  (or as defined)
      Salary Month Day Count → calendar days in the salary month
      Daily Fixed Salary   → Fixed Salary / Salary Month Day Count
      Daily Variable Salary→ Variable Salary / Salary Month Day Count

    Since the uploaded file contains Excel formulas that reference an external
    workbook, we recompute everything from the Total Salary column and the
    calendar days of the report month so the generator is self-contained.

    Returns:
      salary_map : { region_name_lower → {
                        'total_salary'        : float,
                        'fixed_salary'        : float,
                        'variable_salary'     : float,
                        'month_day_count'     : int,
                        'daily_fixed'         : float,
                        'daily_variable'      : float,
                    } }
    """
    log(f"  Loading Region Wise Salary: {os.path.basename(file_path)}")

    df = pd.read_excel(file_path)
    df.columns = df.columns.str.strip()

    # Flexible column matching
    col_map = {c.lower().replace(" ", "_"): c for c in df.columns}

    def find_col(*candidates):
        for c in candidates:
            if c in col_map:
                return col_map[c]
        return None

    region_col = find_col("region")
    total_col  = find_col("total_salary", "totalsalary", "salary")

    if not region_col or not total_col:
        raise ValueError(
            "Region Wise Salary file must have 'Region' and 'Total Salary' columns. "
            f"Found: {list(df.columns)}"
        )

    # Determine salary month day count from the report's end date
    # (number of calendar days in that month)
    import calendar
    month_day_count = calendar.monthrange(report_end_date.year, report_end_date.month)[1]

    salary_map = {}
    for _, row in df.iterrows():
        region = str(row[region_col]).strip()
        if not region or region.lower() == "nan":
            continue
        try:
            total = float(row[total_col])
        except (ValueError, TypeError):
            continue

        fixed    = total * 0.70
        variable = total * 0.30
        daily_fixed    = fixed    / month_day_count
        daily_variable = variable / month_day_count

        salary_map[region.lower()] = {
            "total_salary":    total,
            "fixed_salary":    fixed,
            "variable_salary": variable,
            "month_day_count": month_day_count,
            "daily_fixed":     daily_fixed,
            "daily_variable":  daily_variable,
            "region_display":  region,          # original casing for display
        }

    log(f"  → {len(salary_map)} region salary entries loaded (month days = {month_day_count})")
    return salary_map


def compute_weekly_freq_achievement(activity_df, wallet, daily_target, join_date,
                                     year, month, report_end_date=None):
    """
    Frequency Wise Achievement% — four fixed calendar bands, but with a
    JOIN-DATE-AWARE, DYNAMIC weekly target per band.

    The month is still split into the same four fixed bands as before:
        Week 1 : day  1 –  7
        Week 2 : day  8 – 14
        Week 3 : day 15 – 21
        Week 4 : day 22 – 28   (days 29-31 are not part of any band)

    But each TMR's *effective* range within a band depends on when they
    actually started (join_date = first date they appear in either the
    daily activity or daily summary data):

      • Band entirely BEFORE join_date  → band doesn't exist for this TMR;
        it is skipped and excluded from the average (not counted as 0).
      • Band entirely or partially AFTER join_date → effective range is
        [max(band_start, join_date), band_end].
      • The band end is also capped at report_end_date (if given), so a
        band isn't given credit/target for days beyond the actual data.

    Weekly target for a band = daily_target × (number of NON-FRIDAY days
    in that band's effective range) — this replaces the old fixed
    daily_target × 6.

    Band achievement = unique_agents_visited_in_effective_range / weekly_target
                        (capped at 1.0)
    Final score      = average of only the bands that exist for this TMR.

    Returns float in [0.0, 1.0].
    """
    if activity_df is None or activity_df.empty or not daily_target:
        return 0.0

    tmr = activity_df[activity_df["tmr_wallet"] == wallet]
    if tmr.empty:
        return 0.0

    # Valid visits only (same filters as attendance)
    valid = tmr[
        (tmr["checkin_time"].dt.time  >= BUSINESS_START) &
        (tmr["checkout_time"].dt.time <= BUSINESS_END)  &
        (tmr["visit_seconds"] >= VISIT_MIN_SECONDS)     &
        (tmr["visit_seconds"] <= VISIT_MAX_SECONDS)
    ].copy()

    if valid.empty:
        return 0.0

    valid["agent_wallet"] = valid["agent_wallet"].astype(str).str.strip()
    valid["date"]         = pd.to_datetime(valid["date"]).dt.normalize()

    join_ts = pd.Timestamp(join_date).normalize() if join_date is not None and pd.notna(join_date) else None
    cap_ts  = pd.Timestamp(report_end_date).normalize() if report_end_date is not None and pd.notna(report_end_date) else None

    bands = [(1, 7), (8, 14), (15, 21), (22, 28)]
    achievements = []
    for lo, hi in bands:
        band_start = pd.Timestamp(year=year, month=month, day=lo)
        band_end   = pd.Timestamp(year=year, month=month, day=hi)

        if cap_ts is not None:
            band_end = min(band_end, cap_ts)

        eff_start = max(band_start, join_ts) if join_ts is not None else band_start

        if eff_start > band_end:
            # TMR hadn't joined yet during this band (or band is beyond
            # the data range) → this week doesn't exist for them.
            continue

        band_dates   = pd.date_range(start=eff_start, end=band_end, freq="D")
        working_days = sum(1 for d in band_dates if d.weekday() != 4)   # exclude Fridays

        if working_days <= 0:
            continue

        weekly_target = daily_target * working_days

        band_visits   = valid[(valid["date"] >= eff_start) & (valid["date"] <= band_end)]
        unique_agents = band_visits["agent_wallet"].nunique()

        achievements.append(min(unique_agents / weekly_target, 1.0))

    if not achievements:
        return 0.0

    return round(sum(achievements) / len(achievements), 6)



def load_leave_file(file_path, log):
    """
    Load the TMR National Leave Status file.

    Structure: headers on row 3, data from row 4 onward, column B onward.
    Columns: SL, Wallet Number, TMR Name, DH Code, Leave Type,
             Leave Day, Leave Start Day, Leave End Day, Approver, Remarks

    Rules:
      • Expand each leave row into individual dates (Start → End inclusive).
      • Drop any date that falls on a Friday (Fridays are already holidays).
      • The effective leave day count per row = non-Friday dates in the range.
      • Returns: { normalized_wallet → { date (Timestamp) → True } }
        so callers can check leave per (wallet, date).
    """
    log(f"  Loading Leave file: {os.path.basename(file_path)}")

    # Row 3 = headers, data starts row 4; column B onward (index 1)
    df = pd.read_excel(file_path, header=2)          # 0-indexed: row index 2 = row 3
    df = df.dropna(how="all")

    # Flexible column matching
    df.columns = [str(c).strip() for c in df.columns]
    col_lower  = {c.lower().replace(" ", "_"): c for c in df.columns}

    def fc(*candidates):
        for c in candidates:
            if c in col_lower:
                return col_lower[c]
        return None

    wallet_col = fc("wallet_number", "wallet")
    start_col  = fc("leave_start_day", "start_day", "start")
    end_col    = fc("leave_end_day",   "end_day",   "end")

    if not wallet_col or not start_col or not end_col:
        raise ValueError(
            f"Leave file must have Wallet Number, Leave Start Day, Leave End Day columns. "
            f"Found: {list(df.columns)}"
        )

    leave_map = {}   # { normalized_wallet → set of Timestamps }

    for _, row in df.iterrows():
        wallet = str(row.get(wallet_col, "")).strip()
        if not wallet or wallet.lower() == "nan":
            continue
        wallet = normalize_wallet(wallet)

        start = row.get(start_col)
        end   = row.get(end_col)
        if pd.isna(start) or pd.isna(end):
            continue

        try:
            start = pd.Timestamp(start).normalize()
            end   = pd.Timestamp(end).normalize()
        except Exception:
            continue

        # Expand to individual dates, drop Fridays
        date_range = pd.date_range(start=start, end=end, freq="D")
        valid_days = [d for d in date_range if d.weekday() != 4]  # 4 = Friday

        if not valid_days:
            continue

        if wallet not in leave_map:
            leave_map[wallet] = set()
        leave_map[wallet].update(valid_days)

    total_entries = sum(len(v) for v in leave_map.values())
    log(f"  → {len(leave_map)} TMRs with approved leave | {total_entries} total leave days")
    return leave_map

# =============================================================================
# SECTION 4: ATTENDANCE EVALUATION
# =============================================================================

def evaluate_attendance_from_activity(activity_df, dh_target_map, log):
    """
    For every (tmr_wallet, date) pair in the activity data, decide P or A.

    Logic:
      Step 1 — Filter visits to business hours only (10 AM – 7 PM).
      Step 2 — Further filter to valid duration (3 min ≤ x ≤ 10 min).
      Step 3 — Look up this TMR's daily visit target via their DH code.
      Step 4 — If valid_count ≥ 75% of target → P, else → A.

    Returns:
      attendance_dict  : { (wallet, date) → 'P' or 'A' }
      valid_count_dict : { (wallet, date) → int }   (for debugging / audit)
    """
    attendance   = {}
    valid_counts = {}

    # We need a DH code per TMR to look up targets.
    # Build a TMR→DH map from the activity data (take first occurrence per TMR).
    tmr_dh_map = (
        activity_df[["tmr_wallet", "dh_code"]]
        .dropna(subset=["dh_code"])
        .drop_duplicates("tmr_wallet")
        .set_index("tmr_wallet")["dh_code"]
        .to_dict()
    )

    grouped = activity_df.groupby(["tmr_wallet", "date"])

    for (wallet, date), visits in grouped:

        # ── Step 1: business hours filter ──
        in_hours = visits[
            (visits["checkin_time"].dt.time  >= BUSINESS_START) &
            (visits["checkout_time"].dt.time <= BUSINESS_END)
        ]

        # ── Step 2: duration filter ──
        valid = in_hours[
            (in_hours["visit_seconds"] >= VISIT_MIN_SECONDS) &
            (in_hours["visit_seconds"] <= VISIT_MAX_SECONDS)
        ]

        valid_count = len(valid)
        valid_counts[(wallet, date)] = valid_count

        # ── Step 3: look up target via DH code ──
        dh_code = tmr_dh_map.get(wallet, "")
        target  = dh_target_map.get(dh_code, 30)   # fallback target = 30

        # ── Step 4: compare ──
        attendance[(wallet, date)] = "P" if valid_count >= (target * MIN_VISIT_ACHIEVE) else "A"
        




    return attendance, valid_counts


# =============================================================================
# SECTION 5: REPORT BUILDER
# =============================================================================

def build_report(activity_df, summary_df, dh_target_map, dh_mtype_map, log, leave_map=None):
    """
    Combines all data sources and computes every column in the output report.

    Column logic summary:
      TMR Wallet / Name / Region / DH info  → from summary data
      Daily visit target                     → from DH target file via dh_code
      Weekly agent coverage target           → daily_target × 6  (6-day work week)
      Report Start / End Date                → min/max date across all files
      MTD Work Day                           → non-Friday days in the period
      Friday Count                           → Fridays in the period
      Total Month Days                       → calendar days in the period
      Market Hour Target                     → "8:00:00" for everyone
      Average Work Time                      → total working seconds / days present
                                               (from summary data)
      Average Strike                         → total visits / days with any visits
                                               (from summary data)
      6/1/2026 ... (date columns)            → P / A / blank per day
      Present                                → count of P's
      Approved Leave                         → 0 (requires leave file)
      Payable Day                            → Present count
      Attendance%                            → Present / MTD Work Days
      Market Hour Achievement                → min(avg_work / 8h target, 1.0)
      Strike Rate Achievement%               → min(avg_strike / daily_target, 1.0)

    Returns:
      report_df  : one row per TMR
      all_dates  : sorted list of all Timestamps with data
    """

    # ── Collect every date that appears in either dataset ──
    all_dates = set()
    if not activity_df.empty:
        all_dates.update(activity_df["date"].dropna().unique())
    if not summary_df.empty:
        all_dates.update(summary_df["date_"].dropna().unique())

    all_dates    = sorted(pd.Timestamp(d) for d in all_dates)
    report_start = all_dates[0]
    report_end   = all_dates[-1]

    log(f"\n  Period: {report_start.date()} → {report_end.date()}")
    log(f"  Dates with data: {len(all_dates)}")

    # NOTE: MTD Work Day / Friday Count are now computed PER-TMR (below, inside
    # the row loop) from each TMR's own join date, not from this global range.
    total_month_days = (report_end - report_start).days + 1

    # ── Evaluate attendance from activity data ──
    attendance_dict = {}
    if not activity_df.empty:
        log("  Evaluating attendance (business hours + duration filter)...")
        attendance_dict, _ = evaluate_attendance_from_activity(
            activity_df, dh_target_map, log
        )

    # ── Build summary lookup: (wallet, date) → summary row ──
    summary_lookup = {}
    if not summary_df.empty:
        for _, row in summary_df.iterrows():
            w = row.get("tmr_wallet", "")
            d = row.get("date_")
            if pd.notna(d):
                # If same TMR appears twice for same date (edge case), keep last
                summary_lookup[(str(w), pd.Timestamp(d))] = row

    # ── Track each TMR's first available date across BOTH the daily activity
    #    and daily summary data — this is treated as their "join date" and is
    #    used both for the per-TMR report start and for the dynamic,
    #    join-date-aware weekly frequency-achievement targets below. ──
    tmr_first_summary_date = {}
    if not summary_df.empty:
        for _, row in summary_df.iterrows():
            w = str(row.get("tmr_wallet", "")).strip()
            d = row.get("date_")
            if not w or pd.isna(d):
                continue
            d = pd.Timestamp(d).normalize()
            if w not in tmr_first_summary_date or d < tmr_first_summary_date[w]:
                tmr_first_summary_date[w] = d

    tmr_first_activity_date = {}
    if not activity_df.empty:
        for _, row in activity_df.iterrows():
            w = str(row.get("tmr_wallet", "")).strip()
            d = row.get("date")
            if not w or pd.isna(d):
                continue
            d = pd.Timestamp(d).normalize()
            if w not in tmr_first_activity_date or d < tmr_first_activity_date[w]:
                tmr_first_activity_date[w] = d

    tmr_join_date = {}
    for w in set(tmr_first_summary_date) | set(tmr_first_activity_date):
        candidates = [d for d in (tmr_first_summary_date.get(w), tmr_first_activity_date.get(w)) if d is not None]
        if candidates:
            tmr_join_date[w] = min(candidates)

    # ── Build a TMR→DH map from summary (more complete than activity alone) ──
    tmr_dh_from_summary = {}
    tmr_info_from_summary = {}   # stores name, region, dh_name etc.
    if not summary_df.empty:
        for _, row in summary_df.iterrows():
            w = str(row.get("tmr_wallet", ""))
            if not w:
                continue
            if w not in tmr_dh_from_summary:
                tmr_dh_from_summary[w]   = str(row.get("dh_code", "")).strip()
                tmr_info_from_summary[w] = {
                    "tmr_name":              str(row.get("tmr_name", "")),
                    "region":                str(row.get("region", "")),
                    "dh_code":               str(row.get("dh_code", "")).strip(),
                    "distributor_house_name": str(row.get("distributor_house_name", "")),
                }

    # Also pull TMR→DH from activity for any TMRs not in summary
    tmr_dh_from_activity = {}
    tmr_info_from_activity = {}
    if not activity_df.empty:
        for _, row in activity_df.drop_duplicates("tmr_wallet").iterrows():
            w = str(row.get("tmr_wallet", ""))
            if w not in tmr_dh_from_activity:
                tmr_dh_from_activity[w]   = str(row.get("dh_code", "")).strip()
                tmr_info_from_activity[w] = {
                    "tmr_name":              str(row.get("tmr_name", "")),
                    "region":                str(row.get("region", "")),
                    "dh_code":               str(row.get("dh_code", "")).strip(),
                    "distributor_house_name": str(row.get("distributor_house_name", "")),
                }

    # ── Collect all unique TMR wallets ──
    all_wallets = set()
    all_wallets.update(tmr_info_from_summary.keys())
    all_wallets.update(tmr_info_from_activity.keys())
    all_wallets = sorted(all_wallets)

    log(f"  Total TMRs: {len(all_wallets)}")

    rows = []

    for wallet in all_wallets:

        # ── Resolve TMR metadata (prefer summary, fall back to activity) ──
        info = tmr_info_from_summary.get(wallet) or tmr_info_from_activity.get(wallet, {})
        dh_code = info.get("dh_code", "")

        # ── Look up daily visit target from DH target map ──
        daily_target = dh_target_map.get(dh_code, 30)   # default 30 if DH not found

        # Weekly agent coverage target:
        # 6 working days × daily target (Friday is off, so 6 days in a week)
        weekly_coverage_target = daily_target * 6

        # ── Aggregate working time from summary data ──
        total_work_sec   = 0
        work_days_count  = 0    # days where working_seconds > 0
        total_visits     = 0
        visit_days_count = 0    # days where visit count > 0

        for date in all_dates:
            srow = summary_lookup.get((wallet, date))
            if srow is None:
                continue
            sec = int(srow.get("working_seconds", 0) or 0)
            if sec > 0:
                total_work_sec  += sec
                work_days_count += 1
            vc = int(srow.get("agent_visit_count", 0) or 0)
            if vc > 0:
                total_visits    += vc
                visit_days_count += 1

        avg_work_sec = total_work_sec / work_days_count   if work_days_count  > 0 else 0
        avg_strike   = total_visits   / visit_days_count  if visit_days_count > 0 else 0

        # ── Build per-date attendance columns ──
        date_attendance = {}
        present_count   = 0

        # Build per-wallet leave date set for fast lookup
        wallet_leaves = set()
        if leave_map:
            wallet_leaves = leave_map.get(wallet, set())

        for date in all_dates:

            # Fridays are always blank (holiday)
            if is_friday(date):
                date_attendance[date] = ""
                continue

            # Priority 1: approved leave overrides everything
            if date in wallet_leaves:
                date_attendance[date] = "L"
                continue

            # Priority 2: result from granular activity-based evaluation
            att = attendance_dict.get((wallet, date))

            # Priority 3: if no activity data, fall back to summary visit count
            if att is None:
                srow = summary_lookup.get((wallet, date))
                if srow is not None:
                    vc = int(srow.get("agent_visit_count", 0) or 0)
                    att = "P" if vc >= daily_target * MIN_VISIT_ACHIEVE else "A"
                else:
                    att = ""   # TMR has no data at all for this day

            date_attendance[date] = att
            if att == "P":
                present_count += 1

        # ── Per-TMR report start date (join date) = first available date
        #    across activity + summary data ──
        tmr_report_start = tmr_join_date.get(wallet)
        if tmr_report_start is None or pd.isna(tmr_report_start):
            tmr_report_start = report_start

        # ── Per-TMR working days / Friday count, counted from THIS TMR's
        #    own join date to the report end — NOT the global report range.
        #    A TMR who joined mid-month sees fewer Fridays / working days
        #    than one who was present for the whole period. ──
        tmr_mtd_work_days = count_working_days(tmr_report_start, report_end)
        tmr_friday_count  = count_fridays(tmr_report_start, report_end)

        # ── KPI metrics ──
        mkt_hr_target_sec = MARKET_HOUR_TARGET_H * 3600

        # Market Hour Achievement = actual avg work time / 8h target (cap at 1.0)
        mkt_hr_achieve = min(avg_work_sec / mkt_hr_target_sec, 1.0) if mkt_hr_target_sec else 0

        # Strike Rate Achievement = actual avg visits / daily target (cap at 1.0)
        strike_achieve = min(avg_strike / daily_target, 1.0) if daily_target else 0

        # Attendance %
        att_pct = present_count / tmr_mtd_work_days if tmr_mtd_work_days else 0

        # Payable days = Present (approved leave not tracked here)
        payable = present_count

        # ── Approved leave count (non-Friday leave days in report period) ──
        approved_leave = len([d for d in wallet_leaves if d in set(all_dates)])

        # ── Assemble the row ──
        row_dict = {
            "TMR Wallet":                   format_wallet_for_display(wallet),
            "TMR Name":                     info.get("tmr_name", ""),
            "Region":                       info.get("region", ""),
            "DH Code":                      dh_code,
            "Distributor House Name":       info.get("distributor_house_name", ""),
            "Market type":                  dh_mtype_map.get(dh_code, ""),
            "Daily visit target":           daily_target,
            "Weekly agent coverage target": weekly_coverage_target,
            "Report Start Date":            tmr_report_start.date(),
            "Report till Date":             report_end.date(),
            "MTD Work Day":                 tmr_mtd_work_days,
            "Friday Count":                 tmr_friday_count,
            "Govt Holiday Count":           0,
            "Total Month Days":             total_month_days,
            "Market Hour Target":           seconds_to_hms(mkt_hr_target_sec),
            "Average Work Time":            seconds_to_hms(avg_work_sec),
            "Average Strike":               round(avg_strike, 2),
        }

        # Add one column per date (formatted as M/D/YYYY, e.g. 6/1/2026)
        for date in all_dates:
            # Use %-m and %-d on Linux; %#m and %#d on Windows (both remove zero-padding)
            try:
                label = date.strftime("%-m/%-d/%Y")
            except ValueError:
                label = date.strftime("%#m/%#d/%Y")   # Windows fallback
            row_dict[label] = date_attendance.get(date, "")

        # Summary tail columns
        payable = present_count + approved_leave   # leave days are also payable
        att_pct = present_count / tmr_mtd_work_days if tmr_mtd_work_days else 0
        row_dict["Present"]                  = present_count
        row_dict["Approved Leave"]           = approved_leave
        row_dict["Payable Day"]              = payable
        row_dict["Attendance%"]              = round(att_pct,       4)
        row_dict["Market Hour Achievement"]  = round(mkt_hr_achieve, 4)
        row_dict["Strike Rate Achievement%"] = round(strike_achieve,  4)

        rows.append(row_dict)

    report_df = pd.DataFrame(rows)
    return report_df, all_dates, activity_df


# =============================================================================
# SECTION 6: EXCEL WRITER
# =============================================================================

def write_excel_report(report_df, all_dates, output_path, log,
                       salary_map=None, report_end_date=None, activity_df=None):
    """
    Write the final formatted Excel report.

    Formatting highlights:
      • Blue header row with white bold text (upay brand blue #0054A5)
      • P cells: green background | A cells: red background
      • Friday / no-data cells: gray background
      • Present and Payable Day columns: yellow background (upay #FFD504)
      • Alternating light-blue row tints for readability
      • Freeze panes at column C, row 2 (scroll right for dates, down for TMRs)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TMR Daily Attendance"

    # ── Determine column order ──
    fixed_front = [
        "TMR Wallet", "TMR Name", "Region", "DH Code", "Distributor House Name",
        "Market type", "Daily visit target", "Weekly agent coverage target",
        "Report Start Date", "Report till Date",
        "MTD Work Day", "Friday Count", "Govt Holiday Count", "Total Month Days",
        "Market Hour Target", "Average Work Time", "Average Strike",
    ]
    try:
        date_cols = [d.strftime("%-m/%-d/%Y") for d in all_dates]
    except ValueError:
        date_cols = [d.strftime("%#m/%#d/%Y") for d in all_dates]   # Windows

    fixed_end = [
        "Present", "Approved Leave", "Payable Day",
        "Attendance%", "Market Hour Achievement", "Strike Rate Achievement%",
    ]
    all_cols = fixed_front + date_cols + fixed_end

    # ── Reusable style objects ──
    header_fill  = PatternFill("solid", fgColor=C_BLUE)
    present_fill = PatternFill("solid", fgColor=C_GREEN)
    absent_fill  = PatternFill("solid", fgColor=C_RED)
    gray_fill    = PatternFill("solid", fgColor=C_GRAY)
    yellow_fill  = PatternFill("solid", fgColor=C_YELLOW)
    alt_fill     = PatternFill("solid", fgColor=C_LIGHT)
    white_fill   = PatternFill("solid", fgColor=C_WHITE)

    hdr_font  = Font(name="Arial", bold=True, color=C_WHITE,  size=9)
    data_font = Font(name="Arial",             color="000000", size=9)
    bold_font = Font(name="Arial", bold=True,  color="000000", size=9)
    pw_font   = Font(name="Arial", bold=True,  color="000000", size=9)   # Present/Payable

    thin_side   = Side(border_style="thin",   color="BBBBBB")
    cell_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)

    c_center = Alignment(horizontal="center", vertical="center")
    c_left   = Alignment(horizontal="left",   vertical="center")

    # ── Header row ──
    for ci, col_name in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill      = header_fill
        cell.font      = hdr_font
        cell.alignment = c_center
        cell.border    = cell_border
    ws.row_dimensions[1].height = 28

    # ── Data rows ──
    date_col_set = set(date_cols)   # O(1) lookup

    for ri, (_, row) in enumerate(report_df.iterrows(), start=2):
        base_fill = alt_fill if ri % 2 == 0 else white_fill

        for ci, col_name in enumerate(all_cols, start=1):
            raw = row.get(col_name, "")
            value = "" if pd.isna(raw) else raw

            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border    = cell_border
            cell.font      = data_font
            cell.alignment = c_center

            # ── Cell-specific styling ──
            if col_name in date_col_set:
                # Attendance cells: color-coded P/A/L or gray for blank/Friday
                if value == "P":
                    cell.fill = present_fill
                    cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
                elif value == "A":
                    cell.fill = absent_fill
                    cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
                elif value == "L":
                    cell.fill = PatternFill("solid", fgColor="FF8C00")   # orange = Leave
                    cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
                else:
                    cell.fill = gray_fill

            elif col_name in ("Present", "Payable Day"):
                cell.fill = yellow_fill
                cell.font = pw_font

            elif col_name in ("Attendance%", "Market Hour Achievement", "Strike Rate Achievement%"):
                cell.fill         = base_fill
                cell.number_format = "0.00%"

            else:
                cell.fill = base_fill

            # Left-align text-heavy columns
            if col_name in ("TMR Name", "Distributor House Name", "Region"):
                cell.alignment = c_left

    # ── Freeze panes: first row (header) + first 2 columns locked ──
    ws.freeze_panes = "C2"

    # ── Column widths ──
    for ci, col_name in enumerate(all_cols, start=1):
        col_letter = get_column_letter(ci)
        if col_name in date_col_set:
            ws.column_dimensions[col_letter].width = 8
        elif col_name == "TMR Name":
            ws.column_dimensions[col_letter].width = 24
        elif col_name == "Distributor House Name":
            ws.column_dimensions[col_letter].width = 26
        elif col_name == "TMR Wallet":
            ws.column_dimensions[col_letter].width = 15
        else:
            ws.column_dimensions[col_letter].width = max(len(col_name) + 2, 12)

    # ── Salary Sheet (if salary data was provided) ──
    if salary_map and report_end_date:
        write_salary_sheet(wb, report_df, salary_map, report_end_date, activity_df, log)

    wb.save(output_path)
    log(f"\n  ✓ Report saved → {output_path}")


def write_salary_sheet(wb, report_df, salary_map, report_end_date, activity_df, log):
    """
    Add a 'Salary Sheet' worksheet to the already-open openpyxl Workbook (wb).

    Columns (matching the sample SalarySheet):
      A  TMR Wallet
      B  TMR Name
      C  Region
      D  DH Code
      E  Distributor House Name
      F  Market type
      G  Minimum visit for attendance  (daily_target × 75%)
      H  MTD Work Day
      I  Present Day
      J  Friday Count
      K  Govt Holiday Count
      L  Total Payable Day             = Present + Approved Leave + Govt Holiday
      M  Attendance%                   = min(Present / MTD Work Day, 1)
      N  Daily visit target
      O  Average Strike
      P  Strike Rate Achievement%      = min(Avg Strike / Daily Target, 1)
      Q  Monthly agent coverage target = Daily Target × 6 × weeks-in-month
      R  Frequency Wise Achievement%   = min(total_visits / Monthly Coverage, 1)
      S  Market Hour Target            (H:MM:SS text)
      T  Average Work Time             (H:MM:SS text)
      U  Market Hour Achievement%      = min(Avg Work / Market Hour Target, 1)
      V  Variable KPI Achievement      = P×0.35 + R×0.40 + U×0.25
      W  Fixed salary
      X  Variable salary
      Y  Total salary
      Z  Internet charge               (flat 300)
      AA Cashout Charge                = Total salary × 1%
      AB Final Disbursement Amount     = Total + Internet + Cashout
      AC Remarks
      AD Salary Status

    Row 1: KPI weight labels (P1=0.35, R1=0.40, U1=0.25) + SUBTOTAL summaries
    Row 2: Column headers
    Row 3+: One row per TMR
    """
    log("  Writing Salary Sheet…")

    ws = wb.create_sheet(title="Salary Sheet")

    # ── Style helpers (reuse brand palette already imported in write_excel_report) ──
    header_fill  = PatternFill("solid", fgColor=C_BLUE)
    yellow_fill  = PatternFill("solid", fgColor=C_YELLOW)
    alt_fill     = PatternFill("solid", fgColor=C_LIGHT)
    white_fill   = PatternFill("solid", fgColor=C_WHITE)
    green_fill   = PatternFill("solid", fgColor=C_GREEN)

    hdr_font      = Font(name="Arial", bold=True, color=C_WHITE,  size=9)
    data_font     = Font(name="Arial",             color="000000", size=9)
    bold_font     = Font(name="Arial", bold=True,  color="000000", size=9)
    subtotal_font = Font(name="Arial", bold=True,  color="000000", size=9)

    thin_side   = Side(border_style="thin", color="BBBBBB")
    cell_border = Border(left=thin_side, right=thin_side,
                         top=thin_side,  bottom=thin_side)
    c_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    c_right  = Alignment(horizontal="right",  vertical="center")

    SALARY_COLS = [
        "TMR Wallet", "TMR Name", "Region", "DH Code", "Distributor House Name",
        "Market type", "Minimum_visit_for_attendance",
        "MTD Work Day", "Present Day", "Approved Leave", "Friday Count", "Govt Holiday Count",
        "Total Payable Day", "Attendance%",
        "Daily visit target", "Average Strike", "Strike Rate Achievement%",
        "Monthly agent coverage target", "Frequency Wise Achievement%",
        "Market Hour Target", "Average Work Time", "Market hour Achievement%",
        "Variable KPI Achievement",
        "Fixed salary", "Variable salary", "Total salary",
        "Internet charge", "Cashout Charge", "Final Disbursement Amount",
        "Remarks", "Salary Status",
    ]

    # Column index helpers (1-based)
    col_idx = {name: i + 1 for i, name in enumerate(SALARY_COLS)}

    # ── ROW 1: KPI weight + subtotal row ──
    ws.row_dimensions[1].height = 20

    # KPI weights in the header of their respective columns
    weight_row = {
        col_idx["Strike Rate Achievement%"]:  KPI_WEIGHT_STRIKE,
        col_idx["Frequency Wise Achievement%"]: KPI_WEIGHT_FREQUENCY,
        col_idx["Market hour Achievement%"]:  KPI_WEIGHT_MKT_HOUR,
    }
    # Subtotal labels in the money columns
    subtotal_cols = [
        "Fixed salary", "Variable salary", "Total salary",
        "Internet charge", "Cashout Charge", "Final Disbursement Amount",
    ]
    for ci in range(1, len(SALARY_COLS) + 1):
        cell = ws.cell(row=1, column=ci)
        if ci in weight_row:
            cell.value = weight_row[ci]
            cell.font  = bold_font
            cell.alignment = c_center
        col_name = SALARY_COLS[ci - 1]
        if col_name in subtotal_cols:
            # Will be filled in after data rows are written
            pass
        cell.border = cell_border

    # ── ROW 2: Headers ──
    ws.row_dimensions[2].height = 36
    for ci, col_name in enumerate(SALARY_COLS, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.fill      = header_fill
        cell.font      = hdr_font
        cell.alignment = c_center
        cell.border    = cell_border

    # ── Build a quick lookup from report_df by wallet ──
    # report_df already has all the KPI fields we need
    rdf = report_df.set_index("TMR Wallet")

    # For frequency achievement we need total visits and monthly coverage target.
    # Monthly coverage target = Weekly coverage target / 6 days × total working days
    # Actually it's in report_df as "Weekly agent coverage target" (= daily × 6).
    # The sample uses "Monthly agent coverage target" which we interpret as
    # daily_target × 6 × number_of_weeks in the month  ≈ daily × working_days
    # (the simpler / more common approach: daily × MTD Work Day)
    # We'll use: monthly_coverage = daily_target × mtd_work_days

    data_rows_start = 3

    # Year/month for constructing the four calendar bands used in the
    # frequency-wise achievement calculation (assumes a single-month report
    # period, same assumption the rest of the report already makes).
    report_end_ts = pd.Timestamp(report_end_date) if report_end_date is not None else None
    band_year  = report_end_ts.year  if report_end_ts is not None else None
    band_month = report_end_ts.month if report_end_ts is not None else None

    for ri, (_, row) in enumerate(report_df.iterrows(), start=data_rows_start):
        base_fill = alt_fill if ri % 2 == 0 else white_fill

        wallet   = row["TMR Wallet"]
        region   = str(row.get("Region", ""))
        sal_info = salary_map.get(region.lower(), {})

        # --- pull KPI values from report_df ---
        mtd_work_day   = int(row.get("MTD Work Day", 0) or 0)
        present_day    = int(row.get("Present", 0) or 0)
        friday_count   = int(row.get("Friday Count", 0) or 0)
        govt_holiday   = int(row.get("Govt Holiday Count", 0) or 0)
        daily_target   = int(row.get("Daily visit target", 30) or 30)
        avg_strike     = float(row.get("Average Strike", 0) or 0)
        avg_work_time  = str(row.get("Average Work Time", "0:00:00") or "0:00:00")
        mkt_hr_target  = str(row.get("Market Hour Target", "8:00:00") or "8:00:00")

        # Derived
        approved_leave        = int(row.get("Approved Leave", 0) or 0)
        min_visit_for_att     = int(daily_target * MIN_VISIT_ACHIEVE)
        total_payable_day     = present_day + approved_leave + govt_holiday + friday_count
        attendance_pct        = min(present_day / mtd_work_day, 1.0) if mtd_work_day else 0
        strike_rate_ach       = min(avg_strike / daily_target, 1.0) if daily_target else 0
        weekly_cov_target     = int(row.get("Weekly agent coverage target", daily_target * 6) or daily_target * 6)

        # ── Frequency Wise Achievement (weekly unique agent visits) ──
        # Split month into 4 fixed bands: day 1-7, 8-14, 15-21, 22-28.
        # Each TMR's effective range within a band is trimmed to start at
        # their join date (first date seen in activity/summary data), and
        # the weekly target for that band is daily_target × non-Friday days
        # remaining in the (possibly trimmed) band — not a flat daily×6.
        # Bands that fall entirely before the TMR joined are skipped and
        # excluded from the average. Final score = average of existing bands.
        tmr_join_date = pd.Timestamp(row.get("Report Start Date")) if row.get("Report Start Date") else None
        freq_ach = compute_weekly_freq_achievement(
            activity_df,
            normalize_wallet(wallet),
            daily_target,
            tmr_join_date,
            band_year,
            band_month,
            report_end_date=report_end_ts,
        )

        # ── Market Hour Achievement ──
        # Rule: if avg work time < 8 h target → score = 0
        #       if avg work time ≥ 8 h target → actual ÷ 8 h  (capped at 1)
        avg_work_sec  = hms_to_seconds(avg_work_time)
        mkt_hr_sec    = MARKET_HOUR_TARGET_H * 3600          # 7 h = 25 200 s
        if avg_work_sec < mkt_hr_sec:
            mkt_hr_ach = 0.0
        else:
            mkt_hr_ach = min(avg_work_sec / mkt_hr_sec, 1.0)

        # Variable KPI
        variable_kpi  = (strike_rate_ach * KPI_WEIGHT_STRIKE
                         + freq_ach      * KPI_WEIGHT_FREQUENCY
                         + mkt_hr_ach    * KPI_WEIGHT_MKT_HOUR)

        # Salary calculations
        daily_fixed    = sal_info.get("daily_fixed",    0)
        daily_variable = sal_info.get("daily_variable", 0)
        total_salary   = sal_info.get("total_salary",   0)

        fixed_salary    = total_payable_day * daily_fixed
        variable_salary = variable_kpi * total_payable_day * daily_variable
        total_sal_earned = fixed_salary + variable_salary
        internet_charge  = INTERNET_CHARGE
        cashout_charge   = total_sal_earned * CASHOUT_CHARGE_PCT
        final_disbursement = total_sal_earned + internet_charge + cashout_charge

        market_type = str(row.get("Market type", sal_info.get("region_display", "")))
        # Market type may not be in report_df — try to get from dh_mtype_map via dh_code
        # (it will be passed in if available, otherwise left as empty string)

        values = {
            "TMR Wallet":                   wallet,
            "TMR Name":                     row.get("TMR Name", ""),
            "Region":                       region,
            "DH Code":                      row.get("DH Code", ""),
            "Distributor House Name":       row.get("Distributor House Name", ""),
            "Market type":                  str(row.get("Market type", "") or ""),
            "Minimum_visit_for_attendance": min_visit_for_att,
            "MTD Work Day":                 mtd_work_day,
            "Present Day":                  present_day,
            "Friday Count":                 friday_count,
            "Govt Holiday Count":           govt_holiday,
            "Total Payable Day":            total_payable_day,
            "Approved Leave":               approved_leave,
            "Attendance%":                  round(attendance_pct, 4),
            "Daily visit target":           daily_target,
            "Average Strike":               round(avg_strike, 2),
            "Strike Rate Achievement%":     round(strike_rate_ach, 4),
            "Monthly agent coverage target": weekly_cov_target,
            "Frequency Wise Achievement%":  round(freq_ach, 4),
            "Market Hour Target":           mkt_hr_target,
            "Average Work Time":            avg_work_time,
            "Market hour Achievement%":     round(mkt_hr_ach, 4),
            "Variable KPI Achievement":     round(variable_kpi, 4),
            "Fixed salary":                 round(fixed_salary, 2),
            "Variable salary":              round(variable_salary, 2),
            "Total salary":                 round(total_sal_earned, 2),
            "Internet charge":              internet_charge,
            "Cashout Charge":               round(cashout_charge, 2),
            "Final Disbursement Amount":    round(final_disbursement, 2),
            "Remarks":                      "",
            "Salary Status":                "",
        }

        for ci, col_name in enumerate(SALARY_COLS, start=1):
            val  = values.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = cell_border
            cell.font      = data_font
            cell.alignment = c_center

            # Formatting per column type
            if col_name in ("Attendance%", "Strike Rate Achievement%",
                            "Frequency Wise Achievement%", "Market hour Achievement%",
                            "Variable KPI Achievement"):
                cell.number_format = "0.00%"
                cell.fill = base_fill
            elif col_name in ("Fixed salary", "Variable salary", "Total salary",
                              "Internet charge", "Cashout Charge",
                              "Final Disbursement Amount"):
                cell.number_format = "#,##0.00"
                cell.fill = yellow_fill
                cell.font = bold_font
            elif col_name == "Final Disbursement Amount":
                cell.fill = green_fill
                cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
            elif col_name in ("TMR Name", "Distributor House Name", "Region"):
                cell.fill      = base_fill
                cell.alignment = c_left
            else:
                cell.fill = base_fill

    # ── ROW 1: subtotal values (sum of data column) ──
    last_data_row = data_rows_start + len(report_df) - 1
    for col_name in subtotal_cols:
        ci   = col_idx[col_name]
        col_letter = get_column_letter(ci)
        cell = ws.cell(row=1, column=ci)
        cell.value         = f"=SUM({col_letter}{data_rows_start}:{col_letter}{last_data_row})"
        cell.number_format = "#,##0.00"
        cell.font          = subtotal_font
        cell.alignment     = c_right
        cell.fill          = PatternFill("solid", fgColor=C_YELLOW)
        cell.border        = cell_border

    # ── Column widths ──
    col_widths = {
        "TMR Wallet": 15, "TMR Name": 24, "Region": 14,
        "DH Code": 14, "Distributor House Name": 26, "Market type": 12,
        "Minimum_visit_for_attendance": 14, "MTD Work Day": 10,
        "Present Day": 10, "Approved Leave": 12, "Friday Count": 10, "Govt Holiday Count": 10,
        "Total Payable Day": 10, "Attendance%": 11,
        "Daily visit target": 10, "Average Strike": 10,
        "Strike Rate Achievement%": 14, "Monthly agent coverage target": 14,
        "Frequency Wise Achievement%": 14,
        "Market Hour Target": 12, "Average Work Time": 12,
        "Market hour Achievement%": 14, "Variable KPI Achievement": 14,
        "Fixed salary": 14, "Variable salary": 14, "Total salary": 14,
        "Internet charge": 12, "Cashout Charge": 12,
        "Final Disbursement Amount": 18,
        "Remarks": 16, "Salary Status": 14,
    }
    for ci, col_name in enumerate(SALARY_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(col_name, 12)

    # Freeze panes: column C, row 3 (header + subtotal row locked)
    ws.freeze_panes = "C3"

    log(f"  ✓ Salary Sheet added ({len(report_df)} rows)")


# =============================================================================
# SECTION 7: WINDOWS GUI
# =============================================================================

class TMRReportApp(tk.Tk):
    """
    Main application window.

    Layout:
      ┌────────────────────────────────────────────────┐
      │  upay | TMR Report Generator                   │
      ├────────────────────────────────────────────────┤
      │  [1] Daily Activity Files  [Browse] [Clear]    │
      │      file list box                             │
      │  [2] Datewise Summary Files[Browse] [Clear]    │
      │      file list box                             │
      │  [3] DH Wise Target File   [Browse]            │
      │      single-file label                         │
      │  [4] Output File           [Browse]            │
      │      output path label                         │
      ├────────────────────────────────────────────────┤
      │  [    Generate Report    ]                     │
      ├────────────────────────────────────────────────┤
      │  Log / progress output (scrolled text box)     │
      └────────────────────────────────────────────────┘
    """

    def __init__(self):
        super().__init__()

        self.title("TMR Report Generator — upay Sales Operations")
        self.resizable(True, True)
        self.minsize(750, 680)

        # ── Color palette (matches upay brand) ──
        self.BG       = "#F4F7FC"
        self.BLUE     = "#0054A5"
        self.YELLOW   = "#FFD504"
        self.WHITE    = "#FFFFFF"
        self.TEXT     = "#1A1A2E"
        self.SUBTEXT  = "#5A6A85"
        self.SUCCESS  = "#00873E"
        self.ERROR    = "#C00000"

        self.configure(bg=self.BG)

        # ── State ──
        self.activity_files = []     # list of selected Daily Activity paths
        self.summary_files  = []     # list of selected Datewise Summary paths
        self.dh_target_file  = tk.StringVar()
        self.salary_file     = tk.StringVar()
        self.leave_file      = tk.StringVar()
        self.output_file     = tk.StringVar()

        self._build_ui()
        self._center_window()

    # ── UI builder ────────────────────────────────────────────────────────────

    def _center_window(self):
        """Place the window in the center of the screen on startup."""
        self.update_idletasks()
        w = self.winfo_width()
        h = self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def _build_ui(self):
        """Construct all widgets."""

        # ── Header banner ──
        header = tk.Frame(self, bg=self.BLUE, height=64)
        header.pack(fill="x")
        tk.Label(
            header,
            text="upay  |  TMR Report Generator",
            bg=self.BLUE, fg=self.WHITE,
            font=("Arial", 16, "bold"),
            pady=16,
        ).pack(side="left", padx=24)

        # ── Main content frame ──
        content = tk.Frame(self, bg=self.BG, padx=20, pady=16)
        content.pack(fill="both", expand=True)

        # ── Section 1: Daily Activity Files ──
        self._section_label(content, "1", "TMR Daily Activity Files",
                            "(One file per day — select multiple at once)")

        act_btn_row = tk.Frame(content, bg=self.BG)
        act_btn_row.pack(fill="x", pady=(0, 4))
        self._btn(act_btn_row, "Browse & Add Files", self._browse_activity, primary=True)
        self._btn(act_btn_row, "Clear", self._clear_activity)

        self.act_listbox = self._file_listbox(content, height=4)

        # ── Section 2: Datewise Summary Files ──
        self._section_label(content, "2", "TMR Datewise Summary Files",
                            "(One file per day — select multiple at once)")

        sum_btn_row = tk.Frame(content, bg=self.BG)
        sum_btn_row.pack(fill="x", pady=(0, 4))
        self._btn(sum_btn_row, "Browse & Add Files", self._browse_summary, primary=True)
        self._btn(sum_btn_row, "Clear", self._clear_summary)

        self.sum_listbox = self._file_listbox(content, height=4)

        # ── Section 3: DH Wise Target File ──
        self._section_label(content, "3", "DH Wise Target File",
                            "(Single file — columns: DH Code, Market type, Daily visit target)")

        dh_row = tk.Frame(content, bg=self.BG)
        dh_row.pack(fill="x", pady=(0, 4))
        self._btn(dh_row, "Browse", self._browse_dh, primary=True)
        self.dh_label = tk.Label(
            dh_row, textvariable=self.dh_target_file,
            bg=self.BG, fg=self.SUBTEXT, font=("Arial", 9),
            anchor="w", wraplength=500,
        )
        self.dh_label.pack(side="left", padx=8, fill="x", expand=True)

        # ── Section 4: Region Wise Salary File ──
        self._section_label(content, "4", "Region Wise Salary File",
                            "(Single file — columns: Region, Total Salary)")

        sal_row = tk.Frame(content, bg=self.BG)
        sal_row.pack(fill="x", pady=(0, 4))
        self._btn(sal_row, "Browse", self._browse_salary, primary=True)
        self.sal_label = tk.Label(
            sal_row, textvariable=self.salary_file,
            bg=self.BG, fg=self.SUBTEXT, font=("Arial", 9),
            anchor="w", wraplength=500,
        )
        self.sal_label.pack(side="left", padx=8, fill="x", expand=True)

        # ── Section 5: Leave Status File (optional) ──
        self._section_label(content, "5", "TMR Leave Status File",
                            "(Optional — columns: Wallet Number, Leave Start Day, Leave End Day)")

        leave_row = tk.Frame(content, bg=self.BG)
        leave_row.pack(fill="x", pady=(0, 4))
        self._btn(leave_row, "Browse", self._browse_leave, primary=True)
        self.leave_label = tk.Label(
            leave_row, textvariable=self.leave_file,
            bg=self.BG, fg=self.SUBTEXT, font=("Arial", 9),
            anchor="w", wraplength=500,
        )
        self.leave_label.pack(side="left", padx=8, fill="x", expand=True)

        # ── Section 6: Output File ──
        self._section_label(content, "6", "Output Report File",
                            "(Choose where to save the generated Excel report)")

        out_row = tk.Frame(content, bg=self.BG)
        out_row.pack(fill="x", pady=(0, 12))
        self._btn(out_row, "Browse", self._browse_output, primary=True)
        self.out_label = tk.Label(
            out_row, textvariable=self.output_file,
            bg=self.BG, fg=self.SUBTEXT, font=("Arial", 9),
            anchor="w", wraplength=500,
        )
        self.out_label.pack(side="left", padx=8, fill="x", expand=True)

        # ── Generate button ──
        sep = tk.Frame(content, bg="#CBD5E0", height=1)
        sep.pack(fill="x", pady=(4, 12))

        self.gen_btn = tk.Button(
            content,
            text="⚙  Generate Report",
            command=self._start_generation,
            bg=self.YELLOW, fg=self.TEXT,
            font=("Arial", 12, "bold"),
            relief="flat", cursor="hand2",
            padx=28, pady=10,
            activebackground="#e6bf00",
        )
        self.gen_btn.pack()

        # ── Log panel ──
        log_label = tk.Label(
            content, text="Progress & Log",
            bg=self.BG, fg=self.SUBTEXT,
            font=("Arial", 9, "bold"), anchor="w",
        )
        log_label.pack(fill="x", pady=(16, 2))

        self.log_box = scrolledtext.ScrolledText(
            content, height=10, font=("Consolas", 9),
            bg="#F8FAFF", fg=self.TEXT,
            relief="flat", bd=1,
            state="disabled",
        )
        self.log_box.pack(fill="both", expand=True)

        # Color tags for the log
        self.log_box.tag_config("ok",    foreground=self.SUCCESS)
        self.log_box.tag_config("error", foreground=self.ERROR)
        self.log_box.tag_config("head",  foreground=self.BLUE, font=("Consolas", 9, "bold"))

    # ── Widget helpers ────────────────────────────────────────────────────────

    def _section_label(self, parent, num, title, sub):
        f = tk.Frame(parent, bg=self.BG)
        f.pack(fill="x", pady=(12, 2))
        tk.Label(f, text=f"Step {num}", bg="#E2EAFC", fg=self.BLUE,
                 font=("Arial", 8, "bold"), padx=6, pady=2).pack(side="left")
        tk.Label(f, text=f"  {title}", bg=self.BG, fg=self.TEXT,
                 font=("Arial", 10, "bold")).pack(side="left")
        tk.Label(f, text=f"  {sub}", bg=self.BG, fg=self.SUBTEXT,
                 font=("Arial", 8)).pack(side="left")

    def _btn(self, parent, text, cmd, primary=False):
        bg = self.BLUE if primary else "#CBD5E0"
        fg = self.WHITE if primary else self.TEXT
        b = tk.Button(
            parent, text=text, command=cmd,
            bg=bg, fg=fg,
            font=("Arial", 9, "bold" if primary else "normal"),
            relief="flat", cursor="hand2",
            padx=12, pady=4,
            activebackground="#003f80" if primary else "#b0bec5",
        )
        b.pack(side="left", padx=(0, 6))

    def _file_listbox(self, parent, height=4):
        frame = tk.Frame(parent, bg=self.BG)
        frame.pack(fill="x", pady=(0, 4))
        lb = tk.Listbox(
            frame, height=height,
            font=("Consolas", 8),
            bg="#F0F4FF", fg=self.TEXT,
            selectbackground=self.BLUE, selectforeground=self.WHITE,
            relief="flat", bd=1,
            activestyle="none",
        )
        sb = ttk.Scrollbar(frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=sb.set)
        lb.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        return lb

    # ── Browse callbacks ──────────────────────────────────────────────────────

    def _browse_activity(self):
        """Open a multi-select file dialog for Daily Activity files."""
        files = filedialog.askopenfilenames(
            title="Select TMR Daily Activity Files (multi-select OK)",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        new = [f for f in files if f not in self.activity_files]
        self.activity_files.extend(new)
        self._refresh_listbox(self.act_listbox, self.activity_files)
        self.log(f"Added {len(new)} activity file(s). Total: {len(self.activity_files)}")

    def _browse_summary(self):
        """Open a multi-select file dialog for Datewise Summary files."""
        files = filedialog.askopenfilenames(
            title="Select TMR Datewise Summary Files (multi-select OK)",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        new = [f for f in files if f not in self.summary_files]
        self.summary_files.extend(new)
        self._refresh_listbox(self.sum_listbox, self.summary_files)
        self.log(f"Added {len(new)} summary file(s). Total: {len(self.summary_files)}")

    def _browse_dh(self):
        """Open a single-file dialog for the DH Wise Target file."""
        path = filedialog.askopenfilename(
            title="Select DH Wise Target File",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.dh_target_file.set(path)
            self.log(f"DH target file: {os.path.basename(path)}")

    def _browse_salary(self):
        """Open a single-file dialog for the Region Wise Salary file."""
        path = filedialog.askopenfilename(
            title="Select Region Wise Salary File",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.salary_file.set(path)
            self.log(f"Salary file: {os.path.basename(path)}")

    def _browse_leave(self):
        """Open a single-file dialog for the Leave Status file."""
        path = filedialog.askopenfilename(
            title="Select TMR Leave Status File",
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.leave_file.set(path)
            self.log(f"Leave file: {os.path.basename(path)}")

    def _browse_output(self):
        """Open a save-as dialog to choose the output report path."""
        path = filedialog.asksaveasfilename(
            title="Save Report As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile = f"TMR_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        )
        if path:
            self.output_file.set(path)
            self.log(f"Output: {path}")

    def _clear_activity(self):
        self.activity_files.clear()
        self.act_listbox.delete(0, "end")
        self.log("Activity files cleared.")

    def _clear_summary(self):
        self.summary_files.clear()
        self.sum_listbox.delete(0, "end")
        self.log("Summary files cleared.")

    def _refresh_listbox(self, lb, paths):
        lb.delete(0, "end")
        for p in paths:
            lb.insert("end", os.path.basename(p))

    # ── Logging ──────────────────────────────────────────────────────────────

    def log(self, message, tag=None):
        """Append a message to the log panel (thread-safe via after())."""
        def _write():
            self.log_box.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self.log_box.insert("end", f"[{ts}] {message}\n", tag or "")
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.after(0, _write)

    # ── Report generation ─────────────────────────────────────────────────────

    def _start_generation(self):
        """
        Validate inputs, then kick off the report generation in a background
        thread so the GUI stays responsive (progress appears in the log panel).
        """
        # ── Validation ──
        errors = []
        if not self.activity_files:
            errors.append("• No Daily Activity files selected.")
        if not self.summary_files:
            errors.append("• No Datewise Summary files selected.")
        if not self.dh_target_file.get():
            errors.append("• No DH Wise Target file selected.")
        if not self.output_file.get():
            errors.append("• No output file path chosen.")

        if errors:
            messagebox.showerror("Missing Inputs", "\n".join(errors))
            return

        # ── Disable button during generation ──
        self.gen_btn.config(state="disabled", text="⏳  Generating…")
        self.log("─" * 60, "head")
        self.log("Starting report generation…", "head")

        # Run in background thread to keep UI alive
        t = threading.Thread(target=self._run_generation, daemon=True)
        t.start()

    def _run_generation(self):
        """Background worker — calls all generation functions and logs progress."""
        try:
            log = self.log   # shorthand

            # ── Load files ──
            log("Loading Daily Activity files…")
            activity_df = load_activity_files(self.activity_files, log)

            log("Loading Datewise Summary files…")
            summary_df = load_summary_files(self.summary_files, log)

            log("Loading DH Wise Target file…")
            dh_target_map, dh_mtype_map = load_dh_targets(self.dh_target_file.get(), log)

            # ── Load leave data (optional) ──
            leave_map = None
            if self.leave_file.get():
                log("Loading Leave Status file…")
                leave_map = load_leave_file(self.leave_file.get(), log)

            # ── Build report ──
            log("Building report…")
            report_df, all_dates, activity_df_built = build_report(
                activity_df, summary_df, dh_target_map, dh_mtype_map, log,
                leave_map=leave_map,
            )

            log(f"  → {len(report_df)} TMRs | {len(all_dates)} date columns")

            # ── Load salary data (optional) ──
            salary_map = None
            report_end_date = all_dates[-1] if all_dates else None
            if self.salary_file.get():
                log("Loading Region Wise Salary file…")
                salary_map = load_salary_file(
                    self.salary_file.get(), report_end_date, log
                )

            # ── Write Excel ──
            log("Writing Excel output…")
            write_excel_report(report_df, all_dates, self.output_file.get(), log,
                               salary_map=salary_map, report_end_date=report_end_date,
                               activity_df=activity_df_built)

            log("✓ Done!", "ok")
            log(f"  File: {self.output_file.get()}", "ok")

            # Ask user if they want to open the file
            self.after(0, self._offer_open_file)

        except Exception as exc:
            tb = traceback.format_exc()
            self.log(f"ERROR: {exc}", "error")
            self.log(tb, "error")
            self.after(0, lambda: messagebox.showerror("Generation Failed", str(exc)))

        finally:
            # Re-enable the button on the main thread
            self.after(0, lambda: self.gen_btn.config(
                state="normal", text="⚙  Generate Report"
            ))

    def _offer_open_file(self):
        """After success, ask whether to open the output file."""
        path = self.output_file.get()
        if messagebox.askyesno("Done!", f"Report generated successfully!\n\nOpen file now?\n{path}"):
            os.startfile(path)   # Windows only — opens with default app (Excel)


# =============================================================================
# SECTION 8: ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    app = TMRReportApp()
    app.mainloop()